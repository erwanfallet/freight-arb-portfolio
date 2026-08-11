"""Reader for the user's real Bloomberg export.

Every Bloomberg sheet is a Date/Value block (sometimes several per sheet, side by side).
This module locates the block, checks the expected unit, and returns a clean pd.Series —
never guessing a ticker: the sheet -> series mapping is explicit in `SERIES_SPECS`, built
from actually inspecting the file (not from an assumed name).

No forward-fill here: this is a read, not a modelling step. A gap stays a gap, the same
rule as the freight portfolio's `ingest/contract.py`.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

DEFAULT_PATH = Path("/Users/erwanfallet/Desktop/Data Bloomberg.xlsx")


class BloombergLoaderError(ValueError):
    """Sheet or block not found, or inconsistent read."""


@dataclass(frozen=True)
class SeriesSpec:
    """A Date/Value block identified in the export, with its declared unit.

    `block_index` distinguishes several series within the same sheet (e.g. the "JET"
    sheet has the M1 swap in block 0 and the prompt spot in block 1).

    `scale` corrects a known quoting factor, verified by hand — never guessed. The CBOT
    grain futures (soybean/corn/wheat) are quoted in **cents per bushel** in this export
    (e.g. 1156.50 = 11.565 USD/bu, verified against the real market level); without
    dividing by 100, any formula expecting a price in USD/bu (board_crush_usd_bu,
    financing_cost_usd_t...) comes out a hundred times too large — the kind of error
    that never crashes.

    `valid_from` excludes a period whose economic unit changed over time. USDBRL before
    July 1994 quotes the pre-Plano Real cruzeiro/cruzeiro real (Brazilian
    hyperinflation, values ~0.0004 in 1992): a different currency, not an outlier to
    correct — it's excluded rather than rescaled.
    """

    sheet: str
    block_index: int
    unit: str
    real_id: str
    note: str = ""
    scale: float = 1.0
    valid_from: str | None = None
    valid_to: str | None = None
    min_valid: float | None = None


# Built from actually inspecting the file (an exhaustive dump of the sheets), never from
# an assumed ticker name. See DEMANDE_DONNEES.md and data_coverage.json for full
# traceability.
SERIES_SPECS: dict[str, SeriesSpec] = {
    "jet_swap_m1": SeriesSpec(
        sheet="JET", block_index=0, unit="c/gal BUT UNSTABLE CONVENTION",
        real_id="US Gulf Coast Jet Fuel 54 Grade Swap M1",
        note=(
            "CONFIRMED DATA DEFECT: the series alternates between USD/gal (values ~3-4) "
            "and c/gal (values ~300-400) several times across its history (jumps "
            "detected in 2019-02, 2019-03, 2020-08/09/11, 2023-07, 2026-05). Do not use "
            "as-is for a spread calculation — either normalise via "
            "`normalize_unit_jumps`, or prefer 'jet_spot' which is clean."
        ),
    ),
    "jet_spot": SeriesSpec(
        sheet="JET", block_index=1, unit="c/gal",
        real_id="US Gulf Coast Jet Fuel 54 Prompt Spot",
        note="Verified clean: no convention jump detected over 8542 observations (1990-2026).",
    ),
    "ulsd": SeriesSpec(
        sheet="HO1 Comdty", block_index=0, unit="c/gal",
        real_id="NYMEX ULSD / Heating Oil front month",
        note=(
            "QUOTED IN CENTS PER GALLON — i.e. by VOLUME. The European leg (ice_gasoil) "
            "is in USD per TONNE, i.e. by MASS. The factor between the two is a "
            "density, not a constant: see `freight.chains.products`, where the "
            "uncertainty on this density is worth 91% of the variability of the arb it "
            "is used to compute."
        ),
    ),
    "ice_gasoil": SeriesSpec(
        sheet="QS1 Comdty", block_index=0, unit="USD/t",
        real_id="ICE Gasoil (Low Sulphur) front month",
        note=(
            "European leg of the transatlantic distillate arb, quoted per tonne. Also "
            "used as an MGO proxy in `agri.chains.freight_cf` for lack of a port-by-port "
            "bunker assessment — two uses, one series, and both are documented."
        ),
    ),
    "ttf": SeriesSpec(
        sheet="TZT1 Comdty", block_index=0, unit="EUR/MWh",
        real_id="ICE Endex Dutch TTF Natural Gas Futures front month",
    ),
    "henry_hub": SeriesSpec(
        sheet="NG1 Comdty", block_index=0, unit="USD/mmBtu",
        real_id="NYMEX Henry Hub Natural Gas front month",
    ),
    "eurusd": SeriesSpec(
        sheet="EUR USD", block_index=0, unit="EUR per USD (check the direction)",
        real_id="EURUSD",
    ),
    "brent": SeriesSpec(
        sheet="CO1 Comdty", block_index=0, unit="USD/bbl", real_id="ICE Brent front month",
    ),
    "wti": SeriesSpec(
        sheet="CL1 Comdty", block_index=0, unit="USD/bbl", real_id="NYMEX WTI front month",
    ),
    "dubai": SeriesSpec(
        sheet="Middle East Dubai Crude FOB Fat", block_index=0, unit="USD/bbl",
        real_id="Middle East Dubai Crude FOB Fateh Cargo Spot",
    ),
    "sofr": SeriesSpec(
        sheet="SOFRRATE Index", block_index=0, unit="decimal fraction (0.0433 = 4.33%)",
        real_id="SOFR", scale=0.01,
        note=(
            "DEFECT FOUND AND FIXED: Bloomberg quotes SOFR in PERCENT (5.40 at the peak "
            "of the 2023 tightening), not as a decimal fraction. Added as-is to a "
            "spread already expressed in decimal (250 bps -> 0.025), it produced an "
            "all-in rate of 243% and inflated the financing cost by a factor of ~100. "
            "`scale=0.01` makes the loader's contract uniform: **every rate comes out "
            "as a decimal fraction**, ready to multiply an amount."
        ),
    ),
    # --- ICE softs: cocoa, coffee, sugar. All verified with no convention jump. ---
    "cocoa_ny": SeriesSpec(
        sheet="CC1 Comdty", block_index=0, unit="USD/t", real_id="ICE Cocoa New York front month",
        note="Max observed 12,565 USD/t — matches the real April 2024 peak, confirming the series isn't rescaled.",
    ),
    "cocoa_london": SeriesSpec(
        sheet="QC1 Comdty", block_index=0, unit="GBP/t", real_id="ICE Cocoa London front month",
    ),
    "coffee_arabica": SeriesSpec(
        sheet="KC1 Comdty", block_index=0, unit="c/lb", real_id="ICE Coffee Arabica (C) front month",
    ),
    "coffee_robusta": SeriesSpec(
        sheet="DF1 Comdty", block_index=0, unit="USD/t", real_id="ICE Robusta Coffee front month",
        note=(
            "Identity inferred from its position in the CC1/QC1/KC1/DF1/SB1/QW1 "
            "sequence, NOT confirmed via DES. Level consistency: max 5817 USD/t close "
            "to the real 2024 robusta record (~5800 USD/t) — plausible but to verify "
            "before any binding use."
        ),
    ),
    "sugar_no11": SeriesSpec(
        sheet="SB1 Comdty", block_index=0, unit="c/lb, 96 pol basis", real_id="ICE Sugar No.11 front month",
    ),
    "sugar_no5": SeriesSpec(
        sheet="QW1 Comdty", block_index=0, unit="USD/t", real_id="ICE Sugar No.5 (white) front month",
    ),
    # --- CBOT grains: QUOTED IN CENTS/BUSHEL in this export, scale=0.01 (see SeriesSpec.scale) ---
    "cbot_soybean": SeriesSpec(
        sheet="S 1 Comdty", block_index=0, unit="USD/bu", real_id="CBOT Soybean front month",
        scale=0.01, note="Raw in c/bu (e.g. 1156.50); scale=0.01 -> 11.565 USD/bu.",
    ),
    "cbot_corn": SeriesSpec(
        sheet="C 1 Comdty", block_index=0, unit="USD/bu", real_id="CBOT Corn front month",
        scale=0.01, note="Raw in c/bu; scale=0.01.",
    ),
    "cbot_wheat": SeriesSpec(
        sheet="W 1 Comdty", block_index=0, unit="USD/bu", real_id="CBOT Wheat (SRW) front month",
        scale=0.01, note="Raw in c/bu; scale=0.01.",
    ),
    "cbot_soymeal": SeriesSpec(
        sheet="SM1 Comdty", block_index=0, unit="USD/short ton", real_id="CBOT Soybean Meal front month",
        note="Already in native USD/short ton — no scale (levels 120-550 consistent with no conversion).",
    ),
    "cbot_soyoil": SeriesSpec(
        sheet="BO1 Comdty", block_index=0, unit="c/lb", real_id="CBOT Soybean Oil front month",
        note="Already in native c/lb — no scale (levels 14-90 consistent with no conversion).",
    ),
    "palm_oil_myr": SeriesSpec(
        sheet="KO1 Comdty", block_index=0, unit="MYR/t", real_id="Bursa Malaysia Crude Palm Oil front month",
        note=(
            "QUOTED IN MALAYSIAN RINGGIT, not dollars. Levels 657-8163 over 1995-2026, "
            "consistent with MYR/tonne (palm oil in USD/t is worth 400-1500 over the "
            "same period). The export contains NO USDMYR series at all: any palm-soy "
            "spread would therefore mix two currencies, which is exactly the error this "
            "portfolio tracks. See `oil_substitution`: the only usable window is the "
            "1998-2005 fixed-peg period, where the FX rate is a known constant."
        ),
    ),
    # --- China DCE ---
    "dce_soymeal": SeriesSpec(
        sheet="AE1 Comdty", block_index=0, unit="CNY/t", real_id="DCE Soybean Meal front month",
    ),
    "dce_soyoil": SeriesSpec(
        sheet="SH1 Comdty", block_index=0, unit="CNY/t", real_id="DCE Soybean Oil front month",
    ),
    # --- FX ---
    "usdbrl": SeriesSpec(
        sheet="USDBRL Curncy", block_index=0, unit="BRL per USD", real_id="USDBRL",
        valid_from="1994-07-01",
        note="Before July 1994: pre-Plano Real cruzeiro/cruzeiro real (hyperinflation) — a different currency, excluded rather than rescaled.",
    ),
    "usdcny": SeriesSpec(
        sheet="USDCNY Curncy", block_index=0, unit="CNY per USD", real_id="USDCNY",
    ),
    # --- Dry freight: the P8 route is split across TWO UNIT REGIMES ---
    # Same defect as 'jet_swap_m1', but here it lands on page T1-1's very subject: the
    # confusion between a TCE in USD/DAY and a voyage rate in USD/TONNE.
    "p8_route_usd_t": SeriesSpec(
        sheet="P8 FFA 66kt Santos to Qingdao M", block_index=0, unit="USD/t",
        real_id="P8 FFA 66kt Santos->Qingdao M0, voyage-rate segment",
        valid_from="2021-11-18", min_valid=5.0,
        note=(
            "USD/TONNE SEGMENT ONLY (2021-11-18 -> 2026-08-08, 774 prints, "
            "35.8-84.5 USD/t). Before that date the same cell quotes a TCE in USD/DAY "
            "(2021-07-01 -> 2021-10-30, 103 prints, 24,500-38,000). The two segments "
            "share NO common date and are separated by a 19-day gap: the conversion "
            "factor therefore cannot be calibrated at the junction, since the market "
            "moved between the two. A zero print on 2022-04-30 is dropped by min_valid."
        ),
    ),
    "p8_route_tce_2021": SeriesSpec(
        sheet="P8 FFA 66kt Santos to Qingdao M", block_index=0, unit="USD/day (TCE)",
        real_id="P8 FFA 66kt Santos->Qingdao M0, TCE segment",
        valid_to="2021-10-30",
        note=(
            "The upstream segment of the SAME cell, in USD/day. Kept separate because "
            "a TCE is the voyage model's natural input, while a USD/t rate is its "
            "output — mixing them is exactly the error T1-1 measures."
        ),
    ),
    "bpi": SeriesSpec(
        sheet="BPI Baltic Exchange Panamax Ind", block_index=0, unit="index points",
        real_id="Baltic Panamax Index (average of TC P1A/P2A/P3A/P4)",
        note=(
            "INDEX POINTS, not USD: the BPI is a weighted average of four timecharter "
            "routes, published in points. It doesn't convert to USD/t without going "
            "through the 5TC in USD/day, absent from the export. Used here as cycle "
            "context, never as a term in a cost calculation."
        ),
    ),
    "vlsfo_singapore": SeriesSpec(
        sheet="GX Very Low Sulphur Fuel Oil Bu", block_index=0, unit="USD/t",
        real_id="VLSFO Bunker Singapore (SGSIN) 1800 Prompt",
    ),
}


def _find_date_blocks(rows: list[tuple]) -> list[tuple[int, int]]:
    """Every (row, column) position where a 'Date' cell appears, in the first 12
    rows — a sheet can carry several series side by side."""
    blocks = []
    for i, row in enumerate(rows[:12]):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() in ("date", "dates"):
                blocks.append((i, j))
    return blocks


def load_raw_series(
    key: str, *, path: Path = DEFAULT_PATH, dropna: bool = True
) -> pd.Series:
    """Loads a series by its key in `SERIES_SPECS`. No ffill, no interpolation.

    `dropna=True` (default) drops dates with no value — common at the top of a
    Bloomberg file (e.g. a dated row whose print hasn't landed yet). Pass False to see
    the raw gaps.
    """
    if key not in SERIES_SPECS:
        raise BloombergLoaderError(
            f"unknown key: {key!r}. Available keys: {sorted(SERIES_SPECS)}"
        )
    spec = SERIES_SPECS[key]
    if not path.exists():
        raise BloombergLoaderError(f"file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if spec.sheet not in wb.sheetnames:
        raise BloombergLoaderError(
            f"sheet {spec.sheet!r} missing from the file — has the export changed "
            f"since data_coverage.json? Available sheets: {wb.sheetnames}"
        )

    ws = wb[spec.sheet]
    rows = list(ws.iter_rows(values_only=True))
    blocks = _find_date_blocks(rows)
    if spec.block_index >= len(blocks):
        raise BloombergLoaderError(
            f"block {spec.block_index} requested but sheet {spec.sheet!r} only "
            f"contains {len(blocks)}"
        )
    header_row, col = blocks[spec.block_index]

    dates: list[date] = []
    values: list[float] = []
    for row in rows[header_row + 1 :]:
        if col >= len(row):
            continue
        d = row[col]
        v = row[col + 1] if col + 1 < len(row) else None
        if isinstance(d, (datetime, date)):
            dates.append(pd.Timestamp(d))
            values.append(v)

    series = pd.Series(values, index=pd.DatetimeIndex(dates), name=key).sort_index()
    if spec.scale != 1.0:
        series = series * spec.scale
    if spec.valid_from is not None:
        series = series.loc[pd.Timestamp(spec.valid_from) :]
    if spec.valid_to is not None:
        series = series.loc[: pd.Timestamp(spec.valid_to)]
    if spec.min_valid is not None:
        # A physically impossible print (zero freight, a negative price) is dropped
        # rather than let through: it doesn't represent a market state, and it
        # silently breaks every downstream division.
        series = series[series > spec.min_valid]
    if dropna:
        series = series.dropna()
    return series


def load(key: str, *, path: Path = DEFAULT_PATH) -> pd.Series:
    """Standard shortcut: `load('ttf')`, `load('henry_hub')`, etc."""
    return load_raw_series(key, path=path)


def detect_unit_jumps(series: pd.Series, *, factor_threshold: float = 20.0) -> pd.Series:
    """Detects unit-convention jumps: day over day, a factor > 20x or < 0.05x.

    A real market move never multiplies a price by 100 overnight; a display-convention
    change (USD vs cents) does. This is the check that found the 'jet_swap_m1' defect:
    the series alternates between the two conventions several times across its history.
    Returns the dates where the jump occurs, with the observed ratio — empty if the
    series is clean.
    """
    ratio = series / series.shift(1)
    return ratio[(ratio > factor_threshold) | (ratio < 1.0 / factor_threshold)]


def available_range(key: str, *, path: Path = DEFAULT_PATH) -> tuple[pd.Timestamp, pd.Timestamp, int]:
    """First date, last date, number of non-null observations — for a page's
    diagnostics panel (staleness, history depth)."""
    series = load_raw_series(key, path=path)
    if series.empty:
        raise BloombergLoaderError(f"series {key!r} empty after reading")
    return series.index.min(), series.index.max(), len(series)
